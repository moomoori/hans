import ccxt
import time
import pandas as pd
import pandas_ta as ta
import logging
import json
import os
import sys
from datetime import datetime
import requests
from openpyxl import Workbook, load_workbook

# [로그 설정] 터미널과 파일에 동시 기록
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.FileHandler("output.log"), logging.StreamHandler(sys.stdout)]
)

# ==========================================
# 1. 고정 설정 (API 및 경로)
# ==========================================
BINANCE_API_KEY = 'SmJl6eVqIicyiy1laDllZNQUyLom0QWR88Ic8OnkrsRiOrADdGheXMEyUICwPISy'
BINANCE_SECRET_KEY = 'HlZMTERFJ7vQ3fElmxPO4sgRiAtDSKzupFIJ4DYw71RlMobVyEpUb2eih9HCQGMd'
TELEGRAM_TOKEN = '8449558544:AAH8GXkQ5DiGgY3DYqrkHmVoJjmn9qYcaVI'
CHAT_ID = '6095382920'
CONFIG_FILE = 'config.json'
REPORT_FILE = "trading_report.xlsx"

exchange = ccxt.binance({
    'apiKey': BINANCE_API_KEY,
    'secret': BINANCE_SECRET_KEY,
    'options': {'defaultType': 'future'},
    'enableRateLimit': True
})

# ==========================================
# 2. 유틸리티 및 실시간 설정 함수
# ==========================================

def load_config():
    """뉴스 봇이 수정한 config.json을 실시간으로 읽어옵니다."""
    if not os.path.exists(CONFIG_FILE):
        default_config = {
            "RISK_RATIO": 0.12, 
            "ADX_THRESHOLD": 20, 
            "STOP_LOSS_PCT": 0.015,
            "LEVERAGE": 3
        }
        with open(CONFIG_FILE, 'w') as f:
            json.dump(default_config, f)
        return default_config
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

def log_to_excel(symbol, side, price, funding, balance):
    """매매 내역을 엑셀 파일에 저장합니다."""
    if not os.path.exists(REPORT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["시간", "종목", "포지션", "진입가", "펀딩비(%)", "총자산($)"])
        wb.save(REPORT_FILE)
    wb = load_workbook(REPORT_FILE)
    ws = wb.active
    ws.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), symbol, side, price, f"{funding*100:.4f}%", balance])
    wb.save(REPORT_FILE)

def send_tg(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    try: requests.get(url, params={"chat_id": CHAT_ID, "text": message}, timeout=5)
    except: pass

# ==========================================
# 3. 전략 분석 엔진
# ==========================================

def get_signal(symbol, adx_limit): # adx_limit은 현재 config에서 가져오지만 여기선 사용 안 함
    try:
        # 5분봉 데이터 로드 (거래 빈도를 높이기 위해 15m -> 5m 변경 추천)
        ohlcv = exchange.fetch_ohlcv(symbol, timeframe='5m', limit=50)
        df = pd.DataFrame(ohlcv, columns=['t','o','h','l','c','v'])
        
        # 1. 볼린저 밴드 계산 (표준편차 1.8로 하향하여 더 자주 닿게 설정)
        bb = ta.bbands(df['close'], length=20, std=1.8)
        df['bb_u'] = bb['BBU_20_1.8']
        df['bb_l'] = bb['BBL_20_1.8']
        
        # 2. 이격도(Bias) 계산: 20일 이평선 기준
        df['sma'] = ta.sma(df['close'], length=20)
        df['bias'] = ((df['close'] - df['sma']) / df['sma']) * 100
        
        curr = df.iloc[-1]
        
        # [수정된 진입 조건]
        # LONG: 볼린저 하단 터치(또는 아래) + 이격도가 -0.2% 이하 (과매도)
        if curr['close'] <= curr['bb_l'] and curr['bias'] <= -0.2:
            return "LONG"
        
        # SHORT: 볼린저 상단 터치(또는 위) + 이격도가 0.2% 이상 (과매수)
        if curr['close'] >= curr['bb_u'] and curr['bias'] >= 0.2:
            return "SHORT"
            
        return None
    except Exception as e:
        # logging.error(f"지표 계산 에러 ({symbol}): {e}")
        return None

# ==========================================
# 4. 주문 및 리스크 관리 실행
# ==========================================

def execute_trade(symbol, side, config):
    try:
        # 펀딩비 리스크 필터링
        funding_rate = float(exchange.fetch_funding_rate(symbol)['fundingRate'])
        if (side == "LONG" and funding_rate > 0.0003) or (side == "SHORT" and funding_rate < -0.0003):
            logging.info(f"⏭️ {symbol} 펀딩비 과열로 진입 패스")
            return

        # 자산 확인 및 수량 계산
        balance_info = exchange.fetch_balance()
        total_balance = float(balance_info['info']['totalWalletBalance'])
        
        stake = total_balance * config['RISK_RATIO']
        price = exchange.fetch_ticker(symbol)['last']
        amount = (stake * config['LEVERAGE']) / price
        
        # 레버리지 설정
        exchange.fapiPrivatePostLeverage({"symbol": symbol.replace("/", ""), "leverage": config['LEVERAGE']})

        # 시장가 주문 실행
        if side == "LONG":
            exchange.create_market_buy_order(symbol, amount)
            sl_price = price * (1 - config['STOP_LOSS_PCT'])
            sl_side = 'sell'
        else:
            exchange.create_market_sell_order(symbol, amount)
            sl_price = price * (1 + config['STOP_LOSS_PCT'])
            sl_side = 'buy'

        # 손절가(Stop Loss) 예약
        exchange.create_order(symbol, 'STOP_MARKET', sl_side, amount, None, {'stopPrice': sl_price, 'reduceOnly': True})
        
        # 결과 기록 (엑셀 및 텔레그램)
        log_to_excel(symbol, side, price, funding_rate, total_balance)
        
        msg = (f"🚀 [{side} 진입 완료]\n종목: {symbol}\n진입가: {price}\n"
               f"손절가: {sl_price:.4f}\n리스크비중: {config['RISK_RATIO']*100}%\n"
               f"ADX필터: {config['ADX_THRESHOLD']}")
        send_tg(msg)
        logging.info(f"✅ 주문 성공: {symbol} {side}")

    except Exception as e:
        logging.error(f"❌ 주문 실패 ({symbol}): {e}")

# ==========================================
# 5. 메인 루프 가동
# ==========================================

def run_bot():
    logging.info("🔥 [실시간 동기화 & 엑셀 기록 봇] 가동 시작")
    send_tg("🤖 최종 통합 매매 봇이 시작되었습니다.\n(뉴스 봇 승인 대기 및 엑셀 기록 활성화)")
    
    while True:
        try:
            # 매 루프마다 config.json에서 최신 설정 로드
            current_config = load_config()
            
            markets = exchange.fetch_markets()
            tickers = [m['symbol'] for m in markets if m['active'] and m['linear'] and m['quote'] == 'USDT']
            
            # 거래량 상위 150개 스캔
            for symbol in tickers[:150]:
                signal = get_signal(symbol, current_config['ADX_THRESHOLD'])
                if signal:
                    # 중복 진입 방지 (포지션 체크)
                    pos = exchange.fetch_positions([symbol])
                    if float(pos[0]['info']['positionAmt']) == 0:
                        execute_trade(symbol, signal, current_config)
                        time.sleep(2)

            logging.info(f"🔎 스캔 완료. 현재 설정 [ADX: {current_config['ADX_THRESHOLD']}, RISK: {current_config['RISK_RATIO']}]")
            time.sleep(1)
            
        except Exception as e:
            logging.error(f"메인 루프 에러: {e}")
            time.sleep(10)

if __name__ == "__main__":
    run_bot()
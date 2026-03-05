import requests
import json
import pandas as pd
import time
import os
from datetime import datetime
from bs4 import BeautifulSoup
from collections import Counter
from openpyxl import Workbook, load_workbook

# ==========================================
# 1. 사용자 설정 (반드시 본인 정보로 수정)
# ==========================================
APP_KEY = "PS2EuMD4Vh9SyK8n08lssx8tIGejXJ2qcsRJ"
APP_SECRET = "sR3V5z7tkJSrmGbS27TTg5xjMYa5ILM72b2cg6rYmP0lpEim7QY0aLoOCFaGYYvYPylhqMnKkTLa3w24dHPykETtVuxczvPEl3XKAG3+KqYuos5PN13Nn1UJfCvj6w95/mUmhqdCPd6UIVLyCuAMeNrbRgdcqCbNLW9YwnhOk6urYbOAGqk="
ACC_NO = "68736060"
ACC_CODE = "01" 
TELEGRAM_TOKEN = '8449558544:AAH8GXkQ5DiGgY3DYqrkHmVoJjmn9qYcaVI'
CHAT_ID = '6095382920'

# 매매 기준 및 안전장치 설정
LIMIT_POSITIONS = 5       # 최대 보유 종목 수
STOP_LOSS_RATE = -3.0     # 손절 수익률 (%)
TAKE_PROFIT_RATE = 5.0    # 익절 수익률 (%)
URL_BASE = "https://openapi.koreainvestment.com:9443" # 실전투자용

# 파일 경로
TRADE_LOG_FILE = "stock_trade_journal.xlsx"
ASSET_LOG_FILE = "daily_asset_report.xlsx"

current_positions = {} # {종목코드: 수량}

# ==========================================
# 2. 기본 유틸리티 기능
# ==========================================

def get_access_token():
    url = f"{URL_BASE}/oauth2/tokenP"
    headers = {"content-type": "application/json"}
    body = {"grant_type": "client_credentials", "appkey": APP_KEY, "appsecret": APP_SECRET}
    res = requests.post(url, headers=headers, data=json.dumps(body))
    return res.json()["access_token"]

def send_tg(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    try:
        requests.get(url, params={"chat_id": CHAT_ID, "text": message})
    except: print("텔레그램 전송 실패")

def log_to_excel(file_name, data, headers):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_name)
    wb = load_workbook(file_name)
    ws = wb.active
    ws.append(data)
    wb.save(file_name)

# ==========================================
# 3. 시장 분석 및 종목 추출
# ==========================================

def get_stock_master():
    url = "https://raw.githubusercontent.com/sharebook-kr/stock-master/master/kospi_code.csv"
    try:
        df = pd.read_csv(url)
        return dict(zip(df['name'], df['code'].astype(str).str.zfill(6)))
    except:
        return {"삼성전자": "005930", "SK하이닉스": "000660", "HLB": "028300"}

def get_hot_topics_from_news(master_dict):
    url = "https://news.naver.com/main/list.naver?mode=LS2D&mid=shm&sid1=101&sid2=258"
    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        res = requests.get(url, headers=headers)
        soup = BeautifulSoup(res.text, 'html.parser')
        titles = " ".join([t.get_text().strip() for t in soup.select('.list_body ul li dl dt:not(.photo) a')])
        found = [name for name in master_dict.keys() if name in titles]
        return Counter(found).most_common(5)
    except: return []

# ==========================================
# 4. 주식 조회 및 주문 기능
# ==========================================

def get_stock_stats(token, symbol):
    url = f"{URL_BASE}/uapi/domestic-stock/v1/quotations/inquire-price"
    headers = {
        "Content-Type": "application/json", "authorization": f"Bearer {token}",
        "appkey": APP_KEY, "appsecret": APP_SECRET, "tr_id": "FHKST01010100"
    }
    params = {"fid_cond_mrkt_div_code": "J", "fid_input_iscd": symbol}
    res = requests.get(url, headers=headers, params=params).json()['output']
    return {"price": int(res['stck_prpr']), "change_rate": float(res['prdy_ctrt']), "name": res.get('hts_kor_isnm', symbol)}

def execute_order(token, symbol, qty, is_buy=True):
    tr_id = "TTTC0802U" if is_buy else "TTTC0801U"
    url = f"{URL_BASE}/uapi/domestic-stock/v1/trading/order-cash"
    headers = {
        "Content-Type": "application/json", "authorization": f"Bearer {token}",
        "appkey": APP_KEY, "appsecret": APP_SECRET, "tr_id": tr_id
    }
    body = {
        "CANO": ACC_NO, "ACNT_PRDT_CD": ACC_CODE, "PDNO": symbol,
        "ORD_DVSN": "01", "ORD_QTY": str(qty), "ORD_UNPR": "0"
    }
    requests.post(url, headers=headers, data=json.dumps(body))

# ==========================================
# 5. [신규] 장 마감 전 일괄 전량 매도 기능
# ==========================================

def sell_all_positions(token):
    """보유한 모든 종목을 시장가로 매도"""
    global current_positions
    send_tg("🔔 장 마감 시간이 다가와 모든 종목을 일괄 매도합니다.")
    
    # 최신 잔고 다시 확인
    url = f"{URL_BASE}/uapi/domestic-stock/v1/trading/inquire-balance"
    headers = {"Content-Type":"application/json", "authorization":f"Bearer {token}", "appkey":APP_KEY, "appsecret":APP_SECRET, "tr_id":"TTTC8434R"}
    params = {"CANO":ACC_NO, "ACNT_PRDT_CD":ACC_CODE, "AFHR_FLG":"N", "OVS_CYFD_QI_F":"N", "INQR_DVSN":"02", "UNPR_DVSN":"01", "CTX_AREA_FK100":"", "CTX_AREA_NK100":""}
    
    res = requests.get(url, headers=headers, params=params).json()
    stocks = res.get('output1', [])
    
    for s in stocks:
        symbol, name, qty = s['pdno'], s['prdt_name'], int(s['hldg_qty'])
        if qty > 0:
            execute_order(token, symbol, qty, is_buy=False)
            send_tg(f"📉 [장 마감 전량매도] {name}({symbol}) {qty}주 매도 완료")
            log_to_excel(TRADE_LOG_FILE, [datetime.now(), symbol, name, "장마감매도", "-", qty, s['evlu_pfls_rt']], ["시간", "코드", "명", "구분", "가격", "수량", "수익률"])
    
    current_positions = {}

# ==========================================
# 6. 메인 감시 및 자산 보고 시스템
# ==========================================

def check_balance_and_manage(token):
    global current_positions
    url = f"{URL_BASE}/uapi/domestic-stock/v1/trading/inquire-balance"
    headers = {"Content-Type":"application/json", "authorization":f"Bearer {token}", "appkey":APP_KEY, "appsecret":APP_SECRET, "tr_id":"TTTC8434R"}
    params = {"CANO":ACC_NO, "ACNT_PRDT_CD":ACC_CODE, "AFHR_FLG":"N", "OVS_CYFD_QI_F":"N", "INQR_DVSN":"02", "UNPR_DVSN":"01", "CTX_AREA_FK100":"", "CTX_AREA_NK100":""}
    
    res = requests.get(url, headers=headers, params=params).json()
    stocks = res.get('output1', [])
    output2 = res.get('output2', [{}])[0]
    
    current_positions = {s['pdno']: int(s['hldg_qty']) for s in stocks if int(s['hldg_qty']) > 0}

    for s in stocks:
        symbol, name, qty = s['pdno'], s['prdt_name'], int(s['hldg_qty'])
        profit_rate = float(s['evlu_pfls_rt'])
        if qty <= 0: continue

        if profit_rate <= STOP_LOSS_RATE or profit_rate >= TAKE_PROFIT_RATE:
            type_str = "손절" if profit_rate <= STOP_LOSS_RATE else "익절"
            execute_order(token, symbol, qty, is_buy=False)
            send_tg(f"⚖️ [{type_str} 완료] {name}({symbol}) / 수익률: {profit_rate}%")
            log_to_excel(TRADE_LOG_FILE, [datetime.now(), symbol, name, type_str, "-", qty, profit_rate], ["시간", "코드", "명", "구분", "가격", "수량", "수익률"])

    return output2

# ==========================================
# 7. 메인 실행 루프
# ==========================================

def main():
    token = get_access_token()
    master_dict = get_stock_master()
    last_news_check, last_asset_check = 0, 0
    sold_out_today = False # 당일 청산 여부
    
    send_tg("🤖 주식 단타 봇이 가동되었습니다. (오버나잇 금지 모드)")

    while True:
        try:
            now = datetime.now()
            now_ts = time.time()
            
            # [추가] 장 마감 처리 (오후 3시 10분 ~ 3시 20분 사이 일괄 매도)
            if now.hour == 15 and 10 <= now.minute <= 20:
                if not sold_out_today:
                    sell_all_positions(token)
                    sold_out_today = True
                    send_tg("🏁 당일 매매를 종료합니다. 봇을 정지하거나 내일 아침 다시 시작하세요.")
                time.sleep(60)
                continue

            # 날짜가 바뀌면 청산 여부 초기화
            if now.hour == 9 and now.minute == 0:
                sold_out_today = False

            # 장 운영 시간 외에는 대기 (09:00 ~ 15:10)
            if now.hour < 9 or (now.hour == 15 and now.minute > 10) or now.hour > 15:
                time.sleep(60)
                continue

            # 1. 자산 확인 및 익절/손절 체크
            acc_info = check_balance_and_manage(token)
            
            # 2. 1시간마다 자산 보고
            if now_ts - last_asset_check > 3600:
                total_asset = acc_info.get('tot_evlu_amt', 0)
                cash = acc_info.get('dnca_tot_amt', 0)
                send_tg(f"📊 [자산 리포트] 총액: {format(int(total_asset), ',')}원 / 예수금: {format(int(cash), ',')}원")
                log_to_excel(ASSET_LOG_FILE, [now.strftime('%Y-%m-%d %H:%M:%S'), int(total_asset), int(cash)], ["시간", "총자산", "예수금"])
                last_asset_check = now_ts

            # 3. 30분마다 뉴스 분석 및 매수 전략 실행
            if now_ts - last_news_check > 1800:
                if len(current_positions) < LIMIT_POSITIONS:
                    hot_stocks = get_hot_topics_from_news(master_dict)
                    for name, count in hot_stocks:
                        symbol = master_dict[name]
                        stats = get_stock_stats(token, symbol)
                        if stats['change_rate'] >= 3.0:
                            execute_order(token, symbol, 10, is_buy=True)
                            send_tg(f"🚀 [매수 포착] {name} - 뉴스 & 급등 포착!")
                last_news_check = now_ts

            time.sleep(2)
        except Exception as e:
            print(f"오류: {e}")
            time.sleep(10)

if __name__ == "__main__":
    main()
import ccxt
import time
import pandas as pd
import pandas_ta as ta
import logging
import os
import sys
from datetime import datetime
import requests
from openpyxl import Workbook, load_workbook

# [로그 설정]
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.FileHandler("daytrade_output.log"), logging.StreamHandler(sys.stdout)]
)

# ==========================================
# 1. 고정 설정 (사용자 정보 및 전략 파라미터)
# ==========================================
BINANCE_API_KEY = 'SmJl6eVqIicyiy1laDllZNQUyLom0QWR88Ic8OnkrsRiOrADdGheXMEyUICwPISy'
BINANCE_SECRET_KEY = 'HlZMTERFJ7vQ3fElmxPO4sgRiAtDSKzupFIJ4DYw71RlMobVyEpUb2eih9HCQGMd'
TELEGRAM_TOKEN = '8449558544:AAH8GXkQ5DiGgY3DYqrkHmVoJjmn9qYcaVI'
CHAT_ID = '6095382920'
REPORT_FILE = "daytrade_report.xlsx"

STRATEGY = {
    "TIMEFRAME": "5m",       # 5분봉 기준
    "BB_PERIOD": 20,
    "BB_STD": 1.8,           # 밴드 폭 좁힘 (신호 빈도 증가)
    "BIAS_THRESHOLD": 0.2,   # 이격도 0.2% (민감도 증가)
    "STOP_LOSS_PCT": 0.015,  # 손절 1.5% (여유 있게 수정)
    "RISK_RATIO": 0.2,       # 가용 잔고의 20% 투입 (최소 주문 금액 회피)
    "LEVERAGE": 5            # 레버리지 5배
}

exchange = ccxt.binance({
    'apiKey': BINANCE_API_KEY,
    'secret': BINANCE_SECRET_KEY,
    'options': {'defaultType': 'future'},
    'enableRateLimit': True
})

# ==========================================
# 2. 유틸리티 함수
# ==========================================

def log_to_excel(symbol, side, price, sl_price, balance):
    if not os.path.exists(REPORT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["시간", "종목", "포지션", "진입가", "손절가", "총자산($)"])
        wb.save(REPORT_FILE)
    wb = load_workbook(REPORT_FILE)
    ws = wb.active
    ws.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), symbol, side, price, sl_price, balance])
    wb.save(REPORT_FILE)

def send_tg(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    try: requests.get(url, params={"chat_id": CHAT_ID, "text": message}, timeout=5)
    except: pass

# ==========================================
# 3. 전략 분석 엔진
# ==========================================

def get_signal(symbol):
    try:
        ohlcv = exchange.fetch_ohlcv(symbol, timeframe=STRATEGY["TIMEFRAME"], limit=50)
        df = pd.DataFrame(ohlcv, columns=['t','o','h','l','c','v'])
        
        # 볼린저 밴드 계산
        bb = ta.bbands(df['close'], length=STRATEGY["BB_PERIOD"], std=STRATEGY["BB_STD"])
        df['bb_u'] = bb.iloc[:, 2] # 상단 밴드
        df['bb_l'] = bb.iloc[:, 0] # 하단 밴드
        
        # 이격도 계산
        df['sma'] = ta.sma(df['close'], length=STRATEGY["BB_PERIOD"])
        df['bias'] = ((df['close'] - df['sma']) / df['sma']) * 100
        
        curr = df.iloc[-1]
        
        if curr['close'] <= curr['bb_l'] and curr['bias'] <= -STRATEGY["BIAS_THRESHOLD"]:
            return "LONG"
        if curr['close'] >= curr['bb_u'] and curr['bias'] >= STRATEGY["BIAS_THRESHOLD"]:
            return "SHORT"
            
        return None
    except:
        return None

# ==========================================
# 4. 주문 실행 로직 (정밀화)
# ==========================================

def execute_trade(symbol, side):
    try:
        # 1. 잔고 조회 (선물 계정 내 USDT 가용 잔고)
        balance_info = exchange.fetch_balance()
        free_usdt = float(balance_info.get('USDT', {}).get('free', 0))
        
        if free_usdt < 10:
            logging.warning(f"⚠️ 잔고 부족: {free_usdt} USDT (최소 10 USDT 필요)")
            return

        price = exchange.fetch_ticker(symbol)['last']
        
        # 2. 수량 계산 (자산비중 * 레버리지 / 현재가)
        amount = (free_usdt * STRATEGY["RISK_RATIO"] * STRATEGY["LEVERAGE"]) / price
        
        # 3. 레버리지 설정
        exchange.fapiPrivatePostLeverage({"symbol": symbol.replace("/", ""), "leverage": STRATEGY["LEVERAGE"]})

        # 4. 시장가 주문
        if side == "LONG":
            exchange.create_market_buy_order(symbol, amount)
            sl_price = price * (1 - STRATEGY["STOP_LOSS_PCT"])
            sl_side = 'sell'
        else:
            exchange.create_market_sell_order(symbol, amount)
            sl_price = price * (1 + STRATEGY["STOP_LOSS_PCT"])
            sl_side = 'buy'

        # 5. 스탑로스 설정
        exchange.create_order(symbol, 'STOP_MARKET', sl_side, amount, None, {'stopPrice': sl_price, 'reduceOnly': True})
        
        log_to_excel(symbol, side, price, sl_price, free_usdt)
        send_tg(f"🚀 [{symbol}] {side} 진입\n진입가: {price}\n손절가: {sl_price:.4f}\n가용잔고: {free_usdt:.2f} USDT")
        logging.info(f"✅ {symbol} {side} 주문 성공")

    except Exception as e:
        logging.error(f"❌ 주문 에러 ({symbol}): {e}")

# ==========================================
# 5. 메인 루프 (거래대금 순 스캔)
# ==========================================

def run_bot():
    logging.info("🔥 [바이낸스 선물 상위 거래대금 스캔] 모드 가동")
    send_tg("🤖 봇이 상위 거래량 종목 스캔을 시작합니다.\n(5분봉, 볼린저+이격도 전략)")
    
    while True:
        try:
            # 전체 종목 시세 가져와서 거래대금(quoteVolume) 기준 정렬
            tickers = exchange.fetch_tickers()
            usdt_tickers = [
                {'symbol': s, 'vol': float(v['quoteVolume'])} 
                for s, v in tickers.items() 
                if '/USDT' in s and 'linear' in v.get('info', {}).get('symbol', '').lower() or '/USDT' in s
            ]
            # 거래대금 내림차순 정렬
            sorted_tickers = sorted(usdt_tickers, key=lambda x: x['vol'], reverse=True)
            
            # 상위 50개 종목만 집중 스캔
            for item in sorted_tickers[:50]:
                symbol = item['symbol']
                signal = get_signal(symbol)
                
                if signal:
                    # 포지션 중복 진입 방지
                    pos = exchange.fetch_positions([symbol])
                    if float(pos[0]['info']['positionAmt']) == 0:
                        execute_trade(symbol, signal)
                        time.sleep(2)

            logging.info(f"🔎 상위 50개 종목 스캔 완료... (대기 중)")
            time.sleep(5)
            
        except Exception as e:
            logging.error(f"메인 루프 에러: {e}")
            time.sleep(10)

if __name__ == "__main__":
    run_bot()